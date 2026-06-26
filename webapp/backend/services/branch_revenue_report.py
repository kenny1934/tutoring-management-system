"""Branch revenue report: summer fee collection + July/August regular sessions.

Computes, per secondary branch (MSA/MSB):
- Summer fee collection (receivable = Paid + Fee Sent, collected = Paid),
  broken down by discount tier, with the not-yet-billed application pipeline.
- Regular-course sessions falling in July/August of the summer year, valued
  at the revenue page's per-session rate:
      cost_per_session = (400 * lessons_paid - discount) / lessons_paid
  (migration 030's session_costs view). Published Summer enrollments also
  write session_log rows — those are excluded here, they are already counted
  in the summer receivable.

The same report feeds two consumers:
- GET /summer/revenue/report — JSON summary for the admin in-app view.
- POST /summer/revenue/sheet-refresh — rebuilds the 7-tab workbook and
  replaces the content of the shared Google Sheet (Drive files.update keeps
  the file id / link / title stable).

Sheet auth reuses GOOGLE_SA_KEY_B64 (same service account as the marketing
snapshot). The target sheet must be shared with that SA as Editor; it lives
in a Shared Drive, so every Drive call passes supportsAllDrives=true.
"""
from __future__ import annotations

import base64
import io
import json
import logging
import os
from collections import defaultdict
from datetime import date
from typing import Any

from sqlalchemy.orm import Session

from constants import (
    BASE_FEE_PER_LESSON,
    REGISTRATION_FEE,
    NON_COUNTABLE_STATUS_PATTERNS,
    SessionStatus,
    SummerApplicationStatus,
    hk_now,
    normalize_secondary_location,
)
from models import (
    Discount,
    Enrollment,
    SessionLog,
    Student,
    SummerApplication,
    SummerCourseConfig,
    SummerCourseSlot,
    SummerSession,
)
from utils.summer_discounts import (
    compute_best_discount,
    load_group_context,
    parse_discounts,
)

logger = logging.getLogger(__name__)

BRANCHES = ("MSA", "MSB")
FULL_COURSE_LESSONS = 8

_S = SummerApplicationStatus
PIPELINE_STATUSES = (
    _S.UNDER_REVIEW.value,
    _S.SUBMITTED.value,
    _S.WAITLISTED.value,
    _S.WITHDRAWN.value,
)
# Statuses whose pipeline amounts count toward the "potential" outlook row.
POTENTIAL_STATUSES = (_S.UNDER_REVIEW.value, _S.SUBMITTED.value)


class RevenueSheetConfigError(RuntimeError):
    """Raised when required env vars are missing or Drive auth fails."""


# ---------------------------------------------------------------------------
# Data collection
# ---------------------------------------------------------------------------

def _is_active_session(session_status: str | None) -> bool:
    # Contains-matching (not exact) is deliberate; the patterns + Cancelled
    # cover every non-active make-up origin row. See NON_ACTIVE_SESSION_STATUSES
    # for the exact-match equivalent used elsewhere.
    st = session_status or ""
    return (
        st != SessionStatus.CANCELLED.value
        and not any(p in st for p in NON_COUNTABLE_STATUS_PATTERNS)
    )


def _effective_fee_status(
    app: SummerApplication, enr_payment_by_app: dict[int, str | None]
) -> str:
    """Which collection bucket (Paid / Fee Sent / raw status) an app's fee sits in.

    Publishing flips an application's status to 'Enrolled', so it no longer reads
    as Paid or Fee Sent — without this it would drop out of the receivable totals
    entirely (and isn't a pipeline status either). Post-publish the collection
    state lives on the linked Summer enrollment's payment_status: 'Paid' →
    collected, anything else → still outstanding (Fee Sent). A missing enrollment
    (shouldn't happen — unpublish reverts the status) defaults to Fee Sent so the
    fee stays counted as receivable rather than vanishing.
    """
    if app.application_status == _S.ENROLLED.value:
        return (
            _S.PAID.value
            if enr_payment_by_app.get(app.id) == "Paid"
            else _S.FEE_SENT.value
        )
    return app.application_status


def _collect_branch(
    db: Session,
    config: SummerCourseConfig,
    branch: str,
    app_branch_by_id: dict[int, str | None],
    apps: list[SummerApplication],
    discounts_by_id: dict[int, Discount],
    enr_payment_by_app: dict[int, str | None],
    today: date,
) -> dict[str, Any]:
    """All summer + regular figures and detail rows for one branch."""
    summer_rows: list[dict[str, Any]] = []
    for app in apps:
        if app_branch_by_id.get(app.id) != branch:
            continue
        group_apps, siblings = load_group_context(db, app)
        res = compute_best_discount(app, group_apps, siblings, config, today=today)
        n_lessons = app.lessons_paid or 0
        is_partial = n_lessons < FULL_COURSE_LESSONS
        summer_rows.append({
            "ref": app.reference_code,
            "name": app.student_name,
            "grade": app.grade,
            # Raw status drives display + the pipeline; fee_status is the
            # collection bucket (resolves published 'Enrolled' apps via their
            # enrollment's payment_status).
            "status": app.application_status,
            "fee_status": _effective_fee_status(app, enr_payment_by_app),
            "code": f"PARTIAL-{n_lessons}L" if is_partial else res.code,
            "tier_name": (
                f"Partial booking ({n_lessons} lessons x ${BASE_FEE_PER_LESSON})"
                if is_partial
                else (res.best.name_en if res.best else "No discount")
            ),
            "amount": res.amount,
            "fee": res.final_fee,
            "lessons": app.lessons_paid,
            "paid_at": app.paid_at.date() if app.paid_at else None,
            "buddy": app.buddy_group_id,
        })

    paid = [r for r in summer_rows if r["fee_status"] == _S.PAID.value]
    fee_sent = [r for r in summer_rows if r["fee_status"] == _S.FEE_SENT.value]
    pipeline: dict[str, list[int]] = defaultdict(lambda: [0, 0])
    for r in summer_rows:
        pipeline[r["status"]][0] += 1
        pipeline[r["status"]][1] += r["fee"]
    bycode: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"fee": 0, "name": "", "pn": 0, "pa": 0, "fn": 0, "fa": 0}
    )
    for x in paid + fee_sent:
        b = bycode[x["code"]]
        b["fee"] = x["fee"]
        b["name"] = x["tier_name"]
        if x["fee_status"] == _S.PAID.value:
            b["pn"] += 1
            b["pa"] += x["fee"]
        else:
            b["fn"] += 1
            b["fa"] += x["fee"]

    # Regular sessions in July/August of the summer year.
    year = config.year
    jul1, aug31 = date(year, 7, 1), date(year, 8, 31)
    sessions = (
        db.query(SessionLog)
        .filter(
            SessionLog.session_date >= jul1,
            SessionLog.session_date <= aug31,
            SessionLog.location == branch,
        )
        .all()
    )
    active = [s for s in sessions if _is_active_session(s.session_status)]
    enr_ids = {s.enrollment_id for s in active if s.enrollment_id}
    enrollments: dict[int, Enrollment] = {}
    if enr_ids:
        enrollments = {
            e.id: e
            for e in db.query(Enrollment)
            .filter(
                Enrollment.id.in_(enr_ids),
                Enrollment.enrollment_type != "Summer",
            )
            .all()
        }
    active = [s for s in active if s.enrollment_id in enrollments]
    students: dict[int, Student] = {}
    if enrollments:
        students = {
            st.id: st
            for st in db.query(Student)
            .filter(Student.id.in_({e.student_id for e in enrollments.values()}))
            .all()
        }
    per_enr: dict[int, dict[str, int]] = defaultdict(lambda: {"jul": 0, "aug": 0})
    for s in active:
        per_enr[s.enrollment_id]["jul" if s.session_date.month == 7 else "aug"] += 1

    reg_rows: list[dict[str, Any]] = []
    for eid, e in enrollments.items():
        st = students.get(e.student_id)
        d = discounts_by_id.get(e.discount_id)
        disc_val = int(d.discount_value) if d and d.discount_value else 0
        lessons = e.lessons_paid or 0
        fee = (
            BASE_FEE_PER_LESSON * lessons
            - disc_val
            + (REGISTRATION_FEE if e.is_new_student else 0)
        )
        rate = (BASE_FEE_PER_LESSON * lessons - disc_val) / lessons if lessons else 0
        c = per_enr[eid]
        reg_rows.append({
            "sid": st.school_student_id if st else "",
            "name": st.student_name if st else "",
            "grade": st.grade if st else "",
            "day": e.assigned_day,
            "time": e.assigned_time,
            "lessons": e.lessons_paid,
            "disc": d.discount_name if d else "",
            "disc_val": disc_val,
            "fee": fee,
            "pay": e.payment_status,
            "first": e.first_lesson_date,
            "jul": c["jul"],
            "aug": c["aug"],
            "rate": rate,
            "jul_val": c["jul"] * rate,
            "aug_val": c["aug"] * rate,
        })
    reg_rows.sort(key=lambda r: (r["pay"] or "", r["sid"] or ""))

    term_fees: dict[str, list[int]] = defaultdict(lambda: [0, 0])
    for x in reg_rows:
        term_fees[x["pay"]][0] += 1
        term_fees[x["pay"]][1] += x["fee"]

    return {
        "summer": summer_rows,
        "paid": paid,
        "fee_sent": fee_sent,
        "pipeline": dict(pipeline),
        "bycode": dict(bycode),
        "reg": reg_rows,
        "term_fees": dict(term_fees),
        "jul_n": sum(x["jul"] for x in reg_rows),
        "aug_n": sum(x["aug"] for x in reg_rows),
        "jul_v": sum(x["jul_val"] for x in reg_rows),
        "aug_v": sum(x["aug_val"] for x in reg_rows),
        "enr_jul": sum(1 for x in reg_rows if x["jul"]),
        "enr_aug": sum(1 for x in reg_rows if x["aug"]),
    }


def _branch_of_app(
    app: SummerApplication, slot_locations: dict[int, set[str]]
) -> str | None:
    """Branch by booked slot locations, falling back to preferred location."""
    locs = {
        normalize_secondary_location(loc)
        for loc in slot_locations.get(app.id, set())
        if loc
    }
    if locs:
        return "/".join(sorted(locs)) if len(locs) > 1 else next(iter(locs))
    return normalize_secondary_location(app.preferred_location)


def collect_report_data(db: Session, config: SummerCourseConfig) -> dict[str, Any]:
    """Branch data + tier metadata for both consumers (JSON view, workbook)."""
    today = hk_now().date()
    apps = (
        db.query(SummerApplication)
        .filter(SummerApplication.config_id == config.id)
        .all()
    )
    slot_locations: dict[int, set[str]] = defaultdict(set)
    slot_loc_by_id = {
        s.id: s.location
        for s in db.query(SummerCourseSlot)
        .filter(SummerCourseSlot.config_id == config.id)
        .all()
    }
    for ss in (
        db.query(SummerSession)
        .join(SummerCourseSlot, SummerSession.slot_id == SummerCourseSlot.id)
        .filter(SummerCourseSlot.config_id == config.id)
        .all()
    ):
        loc = slot_loc_by_id.get(ss.slot_id)
        if loc:
            slot_locations[ss.application_id].add(loc)

    app_branch_by_id = {a.id: _branch_of_app(a, slot_locations) for a in apps}
    discounts_by_id = {d.id: d for d in db.query(Discount).all()}

    # Published applications carry status 'Enrolled'; their real collection
    # state lives on the linked Summer enrollment's payment_status. Map it back
    # so those fees stay counted (see _effective_fee_status).
    app_ids = [a.id for a in apps]
    enr_payment_by_app: dict[int, str | None] = {}
    if app_ids:
        enr_payment_by_app = {
            sa_id: pay
            for sa_id, pay in db.query(
                Enrollment.summer_application_id, Enrollment.payment_status
            ).filter(
                Enrollment.enrollment_type == "Summer",
                Enrollment.summer_application_id.in_(app_ids),
            )
        }

    branches = {
        br: _collect_branch(
            db, config, br, app_branch_by_id, apps, discounts_by_id,
            enr_payment_by_app, today,
        )
        for br in BRANCHES
    }
    return {
        "as_of": hk_now(),
        "config": config,
        "tiers": {d.code: d for d in parse_discounts(config)},
        "branches": branches,
    }


# ---------------------------------------------------------------------------
# JSON summary (for the in-app view)
# ---------------------------------------------------------------------------

def _branch_summary(D: dict[str, Any], tiers: dict[str, Any]) -> dict[str, Any]:
    paid_amount = sum(x["fee"] for x in D["paid"])
    fee_sent_amount = sum(x["fee"] for x in D["fee_sent"])
    receivable = paid_amount + fee_sent_amount
    tiers_out = []
    for code, b in sorted(D["bycode"].items(), key=lambda kv: -(kv[1]["pa"] + kv[1]["fa"])):
        tier = tiers.get(code)
        tiers_out.append({
            "code": code,
            "name": b["name"],
            "discount_amount": tier.amount if tier else 0,
            "fee_per_student": b["fee"],
            "paid_count": b["pn"],
            "paid_amount": b["pa"],
            "fee_sent_count": b["fn"],
            "fee_sent_amount": b["fa"],
        })
    pipeline_out = [
        {
            "status": status,
            "students": D["pipeline"][status][0],
            "amount": D["pipeline"][status][1],
        }
        for status in PIPELINE_STATUSES
        if status in D["pipeline"]
    ]
    potential = sum(
        D["pipeline"].get(s, [0, 0])[1] for s in POTENTIAL_STATUSES
    )
    term_fees_out = [
        {"status": ps, "enrollments": n, "amount": amt}
        for ps, (n, amt) in sorted(D["term_fees"].items())
    ]
    return {
        "receivable_students": len(D["paid"]) + len(D["fee_sent"]),
        "receivable_amount": receivable,
        "collected_students": len(D["paid"]),
        "collected_amount": paid_amount,
        "outstanding_students": len(D["fee_sent"]),
        "outstanding_amount": fee_sent_amount,
        "collection_rate_amount": paid_amount / receivable if receivable else 0,
        "collection_rate_students": (
            len(D["paid"]) / (len(D["paid"]) + len(D["fee_sent"]))
            if D["paid"] or D["fee_sent"]
            else 0
        ),
        "tiers": tiers_out,
        "pipeline": pipeline_out,
        "pipeline_potential_amount": potential,
        "regular": {
            "jul_sessions": D["jul_n"],
            "aug_sessions": D["aug_n"],
            "jul_revenue": round(D["jul_v"], 2),
            "aug_revenue": round(D["aug_v"], 2),
            "enrollments_jul": D["enr_jul"],
            "enrollments_aug": D["enr_aug"],
            "term_fees": term_fees_out,
        },
        "outlook_confirmed": round(receivable + D["jul_v"] + D["aug_v"], 2),
        "outlook_with_potential": round(
            receivable + D["jul_v"] + D["aug_v"] + potential, 2
        ),
    }


def report_to_summary(data: dict[str, Any]) -> dict[str, Any]:
    """JSON-friendly summary for GET /summer/revenue/report."""
    tiers = data["tiers"]
    return {
        "as_of": data["as_of"],
        "config_id": data["config"].id,
        "year": data["config"].year,
        "branches": {
            br: _branch_summary(data["branches"][br], tiers) for br in BRANCHES
        },
    }


# ---------------------------------------------------------------------------
# Workbook (7 tabs, mirrors the Google Sheet layout)
# ---------------------------------------------------------------------------

def build_workbook(data: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    NAVY = "1F3864"
    H_FONT = Font(bold=True, color="FFFFFF", size=10)
    H_FILL = PatternFill("solid", fgColor=NAVY)
    S_FONT = Font(bold=True, size=11, color=NAVY)
    B_FONT = Font(size=10)
    BOLD = Font(bold=True, size=10)
    THIN = Border(bottom=Side(style="thin", color="BFBFBF"))
    TOP = Border(top=Side(style="thin", color="404040"))
    CUR = '"$"#,##0'
    CUR2 = '"$"#,##0.00'
    PCT = "0.0%"
    GREEN = PatternFill("solid", fgColor="E2EFDA")

    tiers = data["tiers"]
    year = data["config"].year
    now_str = data["as_of"].strftime("%Y-%m-%d %H:%M") + " (UTC+8)"
    subtitle = (
        f"Summer Course {year} and Regular Course (July to August {year})"
        f"  |  Data as of {now_str}"
    )

    wb = Workbook()

    def header_row(ws, row, headers, widths=None):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = H_FONT
            c.fill = H_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

    def section(ws, row, text):
        ws.cell(row=row, column=1, value=text).font = S_FONT
        return row + 1

    def put(ws, row, col, val, fmt=None, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = BOLD if bold else B_FONT
        if fmt:
            c.number_format = fmt
        return c

    def tier_label(code, name):
        return f"{code} — {name}" + (
            f" (−${tiers[code].amount})" if code in tiers else ""
        )

    # ----- Tab 1: Combined Summary -----
    ws = wb.active
    ws.title = "Combined Summary"
    ws.sheet_view.showGridLines = False
    # B-D also carry currency in the 3-column sections, so they need money width
    for i, w in enumerate([52, 14, 14, 15, 14, 12, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.cell(row=1, column=1, value="MSA + MSB Combined Revenue Summary").font = Font(
        bold=True, size=14, color=NAVY
    )
    ws.cell(row=2, column=1, value=subtitle).font = Font(
        size=9, italic=True, color="595959"
    )

    A, B = data["branches"]["MSA"], data["branches"]["MSB"]
    a_paid = sum(x["fee"] for x in A["paid"])
    b_paid = sum(x["fee"] for x in B["paid"])
    a_rec = a_paid + sum(x["fee"] for x in A["fee_sent"])
    b_rec = b_paid + sum(x["fee"] for x in B["fee_sent"])

    HDR6 = ["", "MSA (n)", "MSA ($)", "MSB (n)", "MSB ($)", "Total (n)", "Total ($)"]

    def row6(ws, rr, label, an, aa, bn, ba, bold=False, fmt=CUR):
        put(ws, rr, 1, label, bold=bold)
        for col, v, f in [
            (2, an, None), (3, aa, fmt), (4, bn, None), (5, ba, fmt),
            (6, an + bn, None), (7, aa + ba, fmt),
        ]:
            put(ws, rr, col, v, f, bold=bold)

    r = 4
    r = section(ws, r, f"1. Summer Course {year} — Fee Collection")
    header_row(ws, r, HDR6)
    row6(ws, r + 1, "Receivable (Paid + Fee Sent)",
         len(A["paid"]) + len(A["fee_sent"]), a_rec,
         len(B["paid"]) + len(B["fee_sent"]), b_rec, bold=True)
    row6(ws, r + 2, "Collected (Paid)", len(A["paid"]), a_paid, len(B["paid"]), b_paid)
    row6(ws, r + 3, "Outstanding (Fee Sent)",
         len(A["fee_sent"]), sum(x["fee"] for x in A["fee_sent"]),
         len(B["fee_sent"]), sum(x["fee"] for x in B["fee_sent"]))
    rr = r + 4
    put(ws, rr, 1, "Collection rate (by amount)", bold=True)
    put(ws, rr, 3, a_paid / a_rec if a_rec else 0, PCT, bold=True)
    put(ws, rr, 5, b_paid / b_rec if b_rec else 0, PCT, bold=True)
    put(ws, rr, 7, (a_paid + b_paid) / (a_rec + b_rec) if a_rec + b_rec else 0,
        PCT, bold=True)
    r = rr + 2

    r = section(ws, r, f"2. Summer Course {year} — Fee Tier Breakdown (Paid + Fee Sent)")
    header_row(ws, r, HDR6)
    EMPTY = {"pn": 0, "pa": 0, "fn": 0, "fa": 0, "name": "", "fee": 0}
    all_codes = sorted(
        set(A["bycode"]) | set(B["bycode"]),
        key=lambda c: -(
            (A["bycode"].get(c, EMPTY)["pa"] + A["bycode"].get(c, EMPTY)["fa"])
            + (B["bycode"].get(c, EMPTY)["pa"] + B["bycode"].get(c, EMPTY)["fa"])
        ),
    )
    rr = r + 1
    for code in all_codes:
        a = A["bycode"].get(code, EMPTY)
        b = B["bycode"].get(code, EMPTY)
        row6(ws, rr, tier_label(code, a["name"] or b["name"]),
             a["pn"] + a["fn"], a["pa"] + a["fa"],
             b["pn"] + b["fn"], b["pa"] + b["fa"])
        rr += 1
    row6(ws, rr, "Total", len(A["paid"]) + len(A["fee_sent"]), a_rec,
         len(B["paid"]) + len(B["fee_sent"]), b_rec, bold=True)
    for col in range(1, 8):
        ws.cell(row=rr, column=col).border = TOP
    r = rr + 2

    r = section(ws, r, f"3. Summer Course {year} — Application Pipeline (not yet billed)")
    header_row(ws, r, HDR6)
    rr = r + 1
    for stt in PIPELINE_STATUSES:
        an, aa = A["pipeline"].get(stt, [0, 0])
        bn, ba = B["pipeline"].get(stt, [0, 0])
        if an + bn == 0:
            continue
        row6(ws, rr, stt, an, aa, bn, ba)
        rr += 1
    r = rr + 1

    r = section(ws, r, f"4. Regular Course — Sessions Falling in July / August {year}")
    header_row(ws, r, ["", "MSA", "MSB", "Total"])
    rows4 = [
        ("July active sessions", A["jul_n"], B["jul_n"], None),
        ("July session revenue (discount-prorated)", A["jul_v"], B["jul_v"], CUR2),
        ("August active sessions", A["aug_n"], B["aug_n"], None),
        ("August session revenue (discount-prorated)", A["aug_v"], B["aug_v"], CUR2),
        ("Enrollments with July sessions", A["enr_jul"], B["enr_jul"], None),
        ("Enrollments with August sessions", A["enr_aug"], B["enr_aug"], None),
    ]
    rr = r + 1
    for label, av, bv, fmt in rows4:
        put(ws, rr, 1, label)
        put(ws, rr, 2, av, fmt)
        put(ws, rr, 3, bv, fmt)
        put(ws, rr, 4, av + bv, fmt)
        rr += 1
    r = rr + 1

    r = section(ws, r, "5. July–August Revenue Outlook (Summer + Regular)")
    header_row(ws, r, ["", "MSA", "MSB", "Total"])
    a_pipe = sum(A["pipeline"].get(s, [0, 0])[1] for s in POTENTIAL_STATUSES)
    b_pipe = sum(B["pipeline"].get(s, [0, 0])[1] for s in POTENTIAL_STATUSES)
    rows5 = [
        ("Summer receivable (Paid + Fee Sent)", float(a_rec), float(b_rec)),
        ("Regular July session revenue", A["jul_v"], B["jul_v"]),
        ("Regular August session revenue", A["aug_v"], B["aug_v"]),
    ]
    rr = r + 1
    for label, av, bv in rows5:
        put(ws, rr, 1, label)
        put(ws, rr, 2, av, CUR2)
        put(ws, rr, 3, bv, CUR2)
        put(ws, rr, 4, av + bv, CUR2)
        rr += 1
    ta = a_rec + A["jul_v"] + A["aug_v"]
    tb = b_rec + B["jul_v"] + B["aug_v"]
    put(ws, rr, 1, "Total outlook (confirmed)", bold=True)
    put(ws, rr, 2, ta, CUR2, bold=True)
    put(ws, rr, 3, tb, CUR2, bold=True)
    put(ws, rr, 4, ta + tb, CUR2, bold=True)
    for col in range(1, 5):
        ws.cell(row=rr, column=col).border = TOP
    rr += 1
    put(ws, rr, 1, "Summer pipeline potential (Under Review + Submitted)")
    put(ws, rr, 2, float(a_pipe), CUR2)
    put(ws, rr, 3, float(b_pipe), CUR2)
    put(ws, rr, 4, float(a_pipe + b_pipe), CUR2)
    rr += 1
    put(ws, rr, 1, "Total outlook (incl. pipeline potential)", bold=True)
    put(ws, rr, 2, ta + a_pipe, CUR2, bold=True)
    put(ws, rr, 3, tb + b_pipe, CUR2, bold=True)
    put(ws, rr, 4, ta + tb + a_pipe + b_pipe, CUR2, bold=True)
    for col in range(1, 5):
        ws.cell(row=rr, column=col).border = TOP

    # ----- Per-branch tabs -----
    for BR in BRANCHES:
        D = data["branches"][BR]
        paid, fee_sent, reg_rows = D["paid"], D["fee_sent"], D["reg"]
        sum_paid = sum(x["fee"] for x in paid)
        sum_fs = sum(x["fee"] for x in fee_sent)
        expected = sum_paid + sum_fs

        ws = wb.create_sheet(f"{BR} Summary")
        ws.sheet_view.showGridLines = False
        for i, w in enumerate([52, 14, 14, 14, 14, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(row=1, column=1, value=f"{BR} Revenue Summary").font = Font(
            bold=True, size=14, color=NAVY
        )
        ws.cell(row=2, column=1, value=subtitle).font = Font(
            size=9, italic=True, color="595959"
        )

        r = 4
        r = section(ws, r, f"1. Summer Course {year} — Fee Collection")
        header_row(ws, r, ["", "Students", "Amount"])
        for i, (label, n, amt) in enumerate([
            ("Receivable (Paid + Fee Sent)", len(paid) + len(fee_sent), expected),
            ("Collected (Paid)", len(paid), sum_paid),
            ("Outstanding (Fee Sent)", len(fee_sent), sum_fs),
        ]):
            put(ws, r + 1 + i, 1, label, bold=(i == 0))
            put(ws, r + 1 + i, 2, n)
            put(ws, r + 1 + i, 3, amt, CUR, bold=(i == 0))
        r += 4
        put(ws, r, 1, "Collection rate (by amount)", bold=True)
        put(ws, r, 3, sum_paid / expected if expected else 0, PCT, bold=True)
        r += 1
        put(ws, r, 1, "Collection rate (by students)")
        put(ws, r, 3,
            len(paid) / (len(paid) + len(fee_sent)) if paid or fee_sent else 0, PCT)
        r += 2

        r = section(ws, r, f"2. Summer Course {year} — Fee Tier Breakdown")
        header_row(ws, r, ["Tier", "Fee / Student", "Paid (n)", "Paid ($)",
                           "Fee Sent (n)", "Fee Sent ($)"])
        rr = r + 1
        for code, b in sorted(D["bycode"].items(),
                              key=lambda kv: -(kv[1]["pa"] + kv[1]["fa"])):
            put(ws, rr, 1, tier_label(code, b["name"]))
            for col, val, fmt in [(2, b["fee"], CUR), (3, b["pn"], None),
                                  (4, b["pa"], CUR), (5, b["fn"], None),
                                  (6, b["fa"], CUR)]:
                put(ws, rr, col, val, fmt)
            rr += 1
        put(ws, rr, 1, "Total", bold=True)
        for col, val, fmt in [(3, len(paid), None), (4, sum_paid, CUR),
                              (5, len(fee_sent), None), (6, sum_fs, CUR)]:
            put(ws, rr, col, val, fmt, bold=True)
        for col in range(1, 7):
            ws.cell(row=rr, column=col).border = TOP
        r = rr + 2

        r = section(ws, r, f"3. Summer Course {year} — Application Pipeline (not yet billed)")
        header_row(ws, r, ["Status", "Students", "Potential Amount"])
        rr = r + 1
        for stt in PIPELINE_STATUSES:
            if stt in D["pipeline"]:
                n, amt = D["pipeline"][stt]
                put(ws, rr, 1, stt)
                put(ws, rr, 2, n)
                put(ws, rr, 3, amt, CUR)
                rr += 1
        r = rr + 1

        r = section(ws, r, f"4. Regular Course — Sessions Falling in July / August {year}")
        header_row(ws, r, ["", "July", "August"])
        rr = r + 1
        for label, a, b, fmt in [
            ("Active sessions (excl. cancelled / pending make-up)",
             D["jul_n"], D["aug_n"], None),
            ("Session revenue (discount-prorated fee per lesson)",
             D["jul_v"], D["aug_v"], CUR2),
            ("Enrollments with sessions in month", D["enr_jul"], D["enr_aug"], None),
        ]:
            put(ws, rr, 1, label)
            put(ws, rr, 2, a, fmt)
            put(ws, rr, 3, b, fmt)
            rr += 1
        rr += 1

        r = section(ws, rr, "5. Regular Course — Term Fees of Enrollments With Jul/Aug Sessions")
        header_row(ws, r, ["Payment Status", "Enrollments", "Term Fees"])
        rr = r + 1
        for ps in sorted(D["term_fees"]):
            n, f = D["term_fees"][ps]
            put(ws, rr, 1, ps)
            put(ws, rr, 2, n)
            put(ws, rr, 3, f, CUR)
            rr += 1
        put(ws, rr, 1, "Total", bold=True)
        put(ws, rr, 2, len(reg_rows), bold=True)
        put(ws, rr, 3, sum(x["fee"] for x in reg_rows), CUR, bold=True)
        for col in range(1, 4):
            ws.cell(row=rr, column=col).border = TOP
        rr += 1
        ws.cell(
            row=rr, column=1,
            value="Term fees span May–July terms; only the Jul/Aug session portion "
                  "(section 4) is attributable to those months.",
        ).font = Font(size=8, italic=True, color="595959")

        # Summer detail tab
        ws2 = wb.create_sheet(f"{BR} Summer Applications")
        header_row(ws2, 1,
                   ["Ref Code", "Student", "Grade", "Status", "Tier Code", "Tier",
                    "Discount", "Fee", "Paid Date", "Buddy Group"],
                   [15, 24, 7, 10, 12, 28, 10, 10, 11, 12])
        ws2.freeze_panes = "A2"
        detail = (sorted(paid, key=lambda x: (x["code"], x["ref"]))
                  + sorted(fee_sent, key=lambda x: (x["code"], x["ref"])))
        rr = 2
        for x in detail:
            vals = [x["ref"], x["name"], x["grade"], x["status"], x["code"],
                    x["tier_name"], -x["amount"] if x["amount"] else 0, x["fee"],
                    x["paid_at"].isoformat() if x["paid_at"] else "", x["buddy"] or ""]
            for col, v in enumerate(vals, 1):
                cc = ws2.cell(row=rr, column=col, value=v)
                cc.font = B_FONT
                cc.border = THIN
                if col in (7, 8):
                    cc.number_format = CUR
            if x["fee_status"] == _S.PAID.value:
                ws2.cell(row=rr, column=4).fill = GREEN
            rr += 1
        put(ws2, rr, 1, "Total", bold=True)
        put(ws2, rr, 8, sum(x["fee"] for x in detail), CUR, bold=True)
        for col in range(1, 11):
            ws2.cell(row=rr, column=col).border = TOP
        ws2.auto_filter.ref = f"A1:J{rr-1}"

        # Regular detail tab
        ws3 = wb.create_sheet(f"{BR} Regular Jul-Aug")
        header_row(ws3, 1,
                   ["Student ID", "Student", "Grade", "Day", "Time", "Lessons Paid",
                    "Discount", "Term Fee", "Payment Status", "First Lesson",
                    "Jul Sessions", "Aug Sessions", "Per-Lesson Rate",
                    "Jul Revenue", "Aug Revenue"],
                   [11, 24, 7, 6, 14, 12, 24, 10, 16, 12, 11, 11, 14, 12, 12])
        ws3.freeze_panes = "A2"
        rr = 2
        for x in reg_rows:
            vals = [x["sid"], x["name"], x["grade"], (x["day"] or "")[:3], x["time"],
                    x["lessons"],
                    x["disc"] + (f" (−${x['disc_val']})" if x["disc_val"] else ""),
                    x["fee"], x["pay"],
                    x["first"].isoformat() if x["first"] else "",
                    x["jul"], x["aug"], x["rate"], x["jul_val"], x["aug_val"]]
            for col, v in enumerate(vals, 1):
                cc = ws3.cell(row=rr, column=col, value=v)
                cc.font = B_FONT
                cc.border = THIN
                if col == 8:
                    cc.number_format = CUR
                if col in (13, 14, 15):
                    cc.number_format = CUR2
            if x["pay"] == "Paid":
                ws3.cell(row=rr, column=9).fill = GREEN
            rr += 1
        put(ws3, rr, 1, "Total", bold=True)
        for col, v, fmt in [(8, sum(x["fee"] for x in reg_rows), CUR),
                            (11, D["jul_n"], None), (12, D["aug_n"], None),
                            (14, D["jul_v"], CUR2), (15, D["aug_v"], CUR2)]:
            put(ws3, rr, col, v, fmt, bold=True)
        for col in range(1, 16):
            ws3.cell(row=rr, column=col).border = TOP
        ws3.auto_filter.ref = f"A1:O{rr-1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Drive upload
# ---------------------------------------------------------------------------

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_drive_service = None


def _get_drive_service():
    """Build (and cache) a Drive v3 service from GOOGLE_SA_KEY_B64.

    Full drive scope (not drive.file): the sheet was created outside this SA,
    it is merely shared with it, so drive.file would not see it.
    """
    global _drive_service
    if _drive_service is not None:
        return _drive_service

    sa_key_b64 = os.environ.get("GOOGLE_SA_KEY_B64")
    if not sa_key_b64:
        raise RevenueSheetConfigError("GOOGLE_SA_KEY_B64 env var is not set")

    from google.oauth2 import service_account as sa_module
    from googleapiclient.discovery import build

    try:
        sa_info = json.loads(base64.b64decode(sa_key_b64))
        creds = sa_module.Credentials.from_service_account_info(
            sa_info, scopes=["https://www.googleapis.com/auth/drive"]
        )
        _drive_service = build("drive", "v3", credentials=creds, cache_discovery=False)
    except RevenueSheetConfigError:
        raise
    except Exception as e:
        raise RevenueSheetConfigError(f"Failed to build Drive client: {e}") from e
    return _drive_service


def push_workbook_to_sheet(xlsx_bytes: bytes, spreadsheet_id: str) -> dict[str, Any]:
    """Replace the Google Sheet's content with the workbook (id/link/title kept).

    Drive converts the uploaded xlsx back into the native spreadsheet. The
    sheet lives in a Shared Drive, hence supportsAllDrives.
    """
    global _drive_service
    from googleapiclient.http import MediaIoBaseUpload

    service = _get_drive_service()
    media = MediaIoBaseUpload(
        io.BytesIO(xlsx_bytes), mimetype=_XLSX_MIME, resumable=False
    )
    try:
        result = (
            service.files()
            .update(
                fileId=spreadsheet_id,
                media_body=media,
                supportsAllDrives=True,
                fields="id,name,modifiedTime",
            )
            .execute()
        )
    except Exception:
        # Cloud Run idle instances drop cached TLS connections; one fresh
        # client retry mirrors google_sheets_service's transient handling.
        _drive_service = None
        service = _get_drive_service()
        result = (
            service.files()
            .update(
                fileId=spreadsheet_id,
                media_body=media,
                supportsAllDrives=True,
                fields="id,name,modifiedTime",
            )
            .execute()
        )
    logger.info(
        "Revenue sheet refreshed: %s (%s)", result.get("name"), result.get("id")
    )
    return result
